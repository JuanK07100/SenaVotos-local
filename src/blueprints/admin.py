from flask import Blueprint, render_template, request, jsonify, render_template_string, send_file
from extensions import db
from models import Usuario, Candidato, Voto, Ficha
from utils.decorators import no_cache
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

bp = Blueprint('admin', __name__, url_prefix='/admin')

@bp.route('/')
@no_cache
def admin():
    votos = db.session.query(
        Usuario.documento,
        Usuario.nombre.label('votante'),
        Candidato.nombre_candidato,
        Usuario.jornada,
        Ficha.idfichas.label('ficha')
    ).join(Voto, Voto.usuarios_idusuario == Usuario.idusuario)\
     .join(Candidato, Voto.candidatos_idcandidato == Candidato.idcandidato)\
     .join(Ficha, Usuario.fichas_idfichas == Ficha.idfichas)\
     .all()
    return render_template('admin.html', votos=votos)

@bp.route('/votos')
def actualizar_votos():
    votos = db.session.query(
        Usuario.documento,
        Usuario.nombre.label('votante'),
        Candidato.nombre_candidato,
        Usuario.jornada,
        Ficha.idfichas.label('ficha')
    ).join(Voto, Voto.usuarios_idusuario == Usuario.idusuario)\
     .join(Candidato, Voto.candidatos_idcandidato == Candidato.idcandidato)\
     .join(Ficha, Usuario.fichas_idfichas == Ficha.idfichas)\
     .all()

    jornadas = sorted(set(v.jornada for v in votos))
    candidatos = sorted(set(v.nombre_candidato for v in votos))

    html_filas = render_template_string("""
        {% for v in votos %}
        <tr>
            <td>{{ v.documento }}</td>
            <td>{{ v.votante }}</td>
            <td>{{ v.ficha or 'Sin ficha' }}</td>
            <td>{{ v.jornada }}</td>
            <td>{{ v.nombre_candidato }}</td>
        </tr>
        {% endfor %}
    """, votos=votos)

    return jsonify({
        "html": html_filas,
        "jornadas": jornadas,
        "candidatos": candidatos,
        "total": len(votos)
    })

@bp.route('/exportar_excel', methods=['POST'])
def exportar_excel():
    jornada_filtro = request.form.get('jornada')
    candidato_filtro = request.form.get('candidato')

    query = db.session.query(
        Usuario.documento,
        Usuario.nombre.label('votante'),
        Usuario.jornada,
        Ficha.idfichas.label('ficha'),
        Candidato.nombre_candidato
    ).join(Voto, Voto.usuarios_idusuario == Usuario.idusuario)\
     .join(Candidato, Voto.candidatos_idcandidato == Candidato.idcandidato)\
     .join(Ficha, Usuario.fichas_idfichas == Ficha.idfichas)

    if jornada_filtro:
        query = query.filter(Usuario.jornada == jornada_filtro)
    if candidato_filtro:
        query = query.filter(Candidato.nombre_candidato == candidato_filtro)

    votos = query.all()

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Votos"

    encabezados = ["Documento del Votante", "Nombre del Votante", "Ficha", "Jornada"]
    header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    ws.append(encabezados)
    for col_num in range(1, 5):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    for v in votos:
        ws.append([v.documento, v.votante, v.ficha, v.jornada])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Hoja Resumen
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Jornada", "Total Votos"])
    for col in range(1, 3):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Conteo por jornada
    jornadas = {"mañana": 0, "tarde": 0, "virtual": 0, "mixta": 0}
    for v in votos:
        if v.jornada in jornadas:
            jornadas[v.jornada] += 1
    row = 2
    for jornada, total in jornadas.items():
        ws2.append([jornada.capitalize(), total])
        for col in range(1, 3):
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    # Conteo por candidato
    ws2.append([])
    ws2.append(["Candidato", "Total Votos"])
    for col in range(1, 3):
        cell = ws2.cell(row=row+1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    candidatos_count = {}
    for v in votos:
        candidatos_count[v.nombre_candidato] = candidatos_count.get(v.nombre_candidato, 0) + 1
    row += 2
    for candidato, total in candidatos_count.items():
        ws2.append([candidato, total])
        for col in range(1, 3):
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    for col in range(1, 3):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # Gráfico de barras (similar al original)
    chart = BarChart()
    chart.type = "col"
    chart.title = "Votos por Jornada"
    chart.y_axis.title = "Votos"
    chart.x_axis.title = "Jornada"
    # Añadir series (simplificado, puedes copiar el original si quieres más detalle)
    data = Reference(ws2, min_col=2, min_row=2, max_row=5)
    categories = Reference(ws2, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)
    ws2.add_chart(chart, "E2")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="lista_votos.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")